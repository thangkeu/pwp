"""XlsxAdapter — trích xuất từ file .xlsx dùng openpyxl. Mỗi sheet → 1 ParsedTable."""

from __future__ import annotations

import io
import zipfile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.adapters.base import ParserAdapter, ParsingError
from app.domain.parsed_document import (
    ParsedDocument,
    ParsedDocumentMetadata,
    ParsedTable,
)


class XlsxAdapter(ParserAdapter):
    @property
    def supported_extensions(self) -> list[str]:
        return ["xlsx", "xlsm"]

    def parse(self, content: bytes, filename: str) -> ParsedDocument:
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except (InvalidFileException, KeyError, OSError, zipfile.BadZipFile) as exc:
            raise ParsingError(f"File XLSX không hợp lệ hoặc hỏng: {filename} ({exc})") from exc

        tables: list[ParsedTable] = []
        text_parts: list[str] = []

        for sheet in workbook.worksheets:
            rows: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                if any(cell.strip() for cell in cells):
                    rows.append(cells)
            if rows:
                tables.append(ParsedTable(rows=rows, caption=sheet.title))
                text_parts.append(f"[Sheet: {sheet.title}]\n" + "\n".join(" | ".join(r) for r in rows))

        text = "\n\n".join(text_parts)
        word_count = len(text.split())

        return ParsedDocument(
            text=text,
            tables=tables,
            metadata=ParsedDocumentMetadata(
                parser_name="XlsxAdapter",
                source_filename=filename,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                page_count=len(workbook.worksheets),
                word_count=word_count,
            ),
        )
